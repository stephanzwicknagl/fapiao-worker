#!/usr/bin/env python3
"""Fill the VAT Reimbursement Claim Form with extracted fapiao data.

Usage:
  Run 1 — write date, fapiao number, quantity (reads from template, saves to OUTPUT):
    python fill_excel.py 1

  Run 2 — write fapiao amount and VAT amount (reads from OUTPUT saved in run 1):
    python fill_excel.py 2
"""

import csv
import datetime
import shutil
import sys
import tomllib
from pathlib import Path

import openpyxl

TEMPLATE = 'VAT Reimbursement Claim Form - January 2026.xlsx'
CSV_FILE = 'fapiaos.csv'
OUTPUT   = 'VAT Reimbursement Claim Form - January 2026 (filled).xlsx'

COL_DATE    = 'B'   # Fapiao date   mm/dd/yyyy
COL_NUMBER  = 'C'   # Fapiao number (text)
COL_CONTENT = 'D'   # Content description (drop-down)
COL_QTY     = 'G'   # Quantity
COL_AMOUNT  = 'I'   # Fapiao amount
COL_VAT     = 'J'   # VAT amount

_MAPPINGS_FILE = Path(__file__).parent.parent / 'mappings.toml'


def _load_mappings() -> dict[str, str]:
    if not _MAPPINGS_FILE.exists():
        return {}
    with open(_MAPPINGS_FILE, 'rb') as f:
        return tomllib.load(f).get('mappings', {})

FIRST_DATA_ROW = 12
MAX_ROWS = 40      # form has rows 12-51


def read_fapiaos(csv_path: str) -> list[dict]:
    with open(csv_path, newline='', encoding='utf-8') as f:
        return list(csv.DictReader(f))


def parse_date(s: str) -> datetime.date | None:
    if not s:
        return None
    try:
        return datetime.date.fromisoformat(s)
    except ValueError:
        return None


def parse_amount(s: str) -> float | None:
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def run1(fapiaos: list[dict], ws) -> None:
    """Write date, fapiao number, quantity, and content description."""
    mappings = _load_mappings()
    print("Run 1: writing date, fapiao number, quantity, content description...")
    for i, row in enumerate(fapiaos):
        r = FIRST_DATA_ROW + i
        date = parse_date(row.get('date', ''))
        number = row.get('fapiao_number') or ''
        seller = row.get('seller') or ''
        category = mappings.get(seller, '') if seller else ''
        ws[f'{COL_DATE}{r}']   = date
        ws[f'{COL_NUMBER}{r}'] = number
        ws[f'{COL_QTY}{r}']    = 1
        if category:
            ws[f'{COL_CONTENT}{r}'] = category
        print(f"  Row {r}: {number}  {date}  qty=1  seller={seller!r}  category={category!r}")


def run2(fapiaos: list[dict], ws) -> None:
    """Write fapiao amount and VAT amount."""
    print("Run 2: writing fapiao amount and VAT amount...")
    for i, row in enumerate(fapiaos):
        r = FIRST_DATA_ROW + i
        amount = parse_amount(row.get('amount', ''))
        vat    = parse_amount(row.get('vat_amount', ''))
        if amount is not None:
            ws[f'{COL_AMOUNT}{r}'] = amount
        if vat is not None:
            ws[f'{COL_VAT}{r}'] = vat
        print(f"  Row {r}: ¥{amount}  VAT ¥{vat}")


def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ('1', '2'):
        print(__doc__)
        sys.exit(1)

    run = int(sys.argv[1])

    if not Path(CSV_FILE).exists():
        print(f"Error: {CSV_FILE} not found. Run extract_fapiaos.py first.")
        sys.exit(1)

    fapiaos = read_fapiaos(CSV_FILE)
    if len(fapiaos) > MAX_ROWS:
        print(f"Warning: {len(fapiaos)} fapiaos but form only has {MAX_ROWS} rows. Truncating.")
        fapiaos = fapiaos[:MAX_ROWS]
    print(f"Read {len(fapiaos)} fapiaos from {CSV_FILE}")

    if run == 1:
        # Start from the original template
        if not Path(TEMPLATE).exists():
            print(f"Error: {TEMPLATE} not found.")
            sys.exit(1)
        shutil.copy2(TEMPLATE, OUTPUT)
        source = OUTPUT
    else:
        # Continue from the file saved by run 1
        if not Path(OUTPUT).exists():
            print(f"Error: {OUTPUT} not found. Run with argument 1 first.")
            sys.exit(1)
        source = OUTPUT

    wb = openpyxl.load_workbook(source, keep_vba=False)
    ws = wb.active

    if run == 1:
        run1(fapiaos, ws)
    else:
        run2(fapiaos, ws)

    wb.save(OUTPUT)
    print(f"\nSaved to: {OUTPUT}")


if __name__ == '__main__':
    main()
