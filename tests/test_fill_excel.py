"""Tests for fapiao/fill.py — two-pass filling logic, mappings lookup, row limits."""

import datetime

import openpyxl
import pytest

import fapiao.fill as fill_excel
from fapiao.fill import (
    FIRST_DATA_ROW,
    MAX_ROWS,
    parse_amount,
    parse_date,
    read_fapiaos,
    run1,
    run2,
)

# ── helpers ───────────────────────────────────────────────────────────────────

def test_parse_date_valid():
    assert parse_date('2024-03-15') == datetime.date(2024, 3, 15)


def test_parse_date_empty():
    assert parse_date('') is None
    assert parse_date(None) is None


def test_parse_date_invalid():
    assert parse_date('not-a-date') is None


def test_parse_amount_valid():
    assert parse_amount('188.50') == pytest.approx(188.50)
    assert parse_amount('0.00') == pytest.approx(0.0)


def test_parse_amount_empty():
    assert parse_amount('') is None
    assert parse_amount(None) is None


def test_parse_amount_invalid():
    assert parse_amount('abc') is None


# ── read_fapiaos ──────────────────────────────────────────────────────────────

def test_read_fapiaos(tmp_path):
    csv_content = (
        'source_file,page,fapiao_number,date,amount,vat_amount,seller\n'
        'test.pdf,1,012345678901234,2024-03-15,188.50,12.33,沃尔玛\n'
    )
    p = tmp_path / 'fapiaos.csv'
    p.write_text(csv_content, encoding='utf-8')
    rows = read_fapiaos(str(p))
    assert len(rows) == 1
    assert rows[0]['fapiao_number'] == '012345678901234'
    assert rows[0]['amount'] == '188.50'


# ── run1 ──────────────────────────────────────────────────────────────────────

def _make_ws():
    wb = openpyxl.Workbook()
    return wb.active


def _fapiao(**kwargs):
    defaults = {
        'fapiao_number': '012345678901234',
        'date': '2024-03-15',
        'amount': '188.50',
        'vat_amount': '12.33',
        'seller': '沃尔玛（湖北）商业零售有限公司',
    }
    defaults.update(kwargs)
    return defaults


def test_run1_writes_date_and_number(tmp_path, monkeypatch):
    mappings = {}
    monkeypatch.setattr(fill_excel, '_load_mappings', lambda: mappings)

    ws = _make_ws()
    run1([_fapiao()], ws)

    row = FIRST_DATA_ROW
    assert ws[f'B{row}'].value == datetime.date(2024, 3, 15)
    assert ws[f'C{row}'].value == '012345678901234'
    assert ws[f'G{row}'].value == 1


def test_run1_writes_category_from_mappings(tmp_path, monkeypatch):
    mappings = {'沃尔玛（湖北）商业零售有限公司': 'Groceries'}
    monkeypatch.setattr(fill_excel, '_load_mappings', lambda: mappings)

    ws = _make_ws()
    run1([_fapiao()], ws)

    row = FIRST_DATA_ROW
    assert ws[f'D{row}'].value == 'Groceries'


def test_run1_skips_category_when_not_mapped(monkeypatch):
    monkeypatch.setattr(fill_excel, '_load_mappings', lambda: {})

    ws = _make_ws()
    run1([_fapiao()], ws)

    row = FIRST_DATA_ROW
    assert ws[f'D{row}'].value is None


def test_run1_skips_category_when_empty_mapping(monkeypatch):
    mappings = {'沃尔玛（湖北）商业零售有限公司': ''}
    monkeypatch.setattr(fill_excel, '_load_mappings', lambda: mappings)

    ws = _make_ws()
    run1([_fapiao()], ws)

    row = FIRST_DATA_ROW
    assert ws[f'D{row}'].value is None


def test_run1_handles_missing_date(monkeypatch):
    monkeypatch.setattr(fill_excel, '_load_mappings', lambda: {})
    ws = _make_ws()
    run1([_fapiao(date='')], ws)
    assert ws[f'B{FIRST_DATA_ROW}'].value is None


def test_run1_multiple_rows_use_correct_offsets(monkeypatch):
    monkeypatch.setattr(fill_excel, '_load_mappings', lambda: {})
    ws = _make_ws()
    fapiaos = [
        _fapiao(fapiao_number='000000000000001', date='2024-01-01'),
        _fapiao(fapiao_number='000000000000002', date='2024-01-02'),
        _fapiao(fapiao_number='000000000000003', date='2024-01-03'),
    ]
    run1(fapiaos, ws)
    assert ws[f'C{FIRST_DATA_ROW}'].value == '000000000000001'
    assert ws[f'C{FIRST_DATA_ROW + 1}'].value == '000000000000002'
    assert ws[f'C{FIRST_DATA_ROW + 2}'].value == '000000000000003'


# ── run2 ──────────────────────────────────────────────────────────────────────

def test_run2_writes_amount_and_vat():
    ws = _make_ws()
    run2([_fapiao()], ws)

    row = FIRST_DATA_ROW
    assert ws[f'I{row}'].value == pytest.approx(188.50)
    assert ws[f'J{row}'].value == pytest.approx(12.33)


def test_run2_skips_none_amount():
    ws = _make_ws()
    run2([_fapiao(amount='', vat_amount='')], ws)

    row = FIRST_DATA_ROW
    assert ws[f'I{row}'].value is None
    assert ws[f'J{row}'].value is None


def test_run2_multiple_rows(monkeypatch):
    ws = _make_ws()
    fapiaos = [
        _fapiao(amount='100.00', vat_amount='9.00'),
        _fapiao(amount='200.00', vat_amount='18.00'),
    ]
    run2(fapiaos, ws)
    assert ws[f'I{FIRST_DATA_ROW}'].value == pytest.approx(100.00)
    assert ws[f'I{FIRST_DATA_ROW + 1}'].value == pytest.approx(200.00)
    assert ws[f'J{FIRST_DATA_ROW + 1}'].value == pytest.approx(18.00)


# ── row limit ─────────────────────────────────────────────────────────────────

def test_max_rows_constant():
    assert MAX_ROWS == 40


def test_run1_respects_max_rows(monkeypatch):
    """run1 itself doesn't truncate — that's the caller's job; verify it fills all provided rows."""
    monkeypatch.setattr(fill_excel, '_load_mappings', lambda: {})
    ws = _make_ws()
    fapiaos = [_fapiao(fapiao_number=f'{i:015d}') for i in range(MAX_ROWS)]
    run1(fapiaos, ws)
    # Last row should be written
    assert ws[f'C{FIRST_DATA_ROW + MAX_ROWS - 1}'].value == f'{MAX_ROWS - 1:015d}'
    # Row beyond MAX should be empty
    assert ws[f'C{FIRST_DATA_ROW + MAX_ROWS}'].value is None


# ── _load_mappings ─────────────────────────────────────────────────────────────

def test_load_mappings_returns_empty_when_file_missing(tmp_path, monkeypatch):
    monkeypatch.setattr(fill_excel, '_MAPPINGS_FILE', tmp_path / 'nonexistent.toml')
    result = fill_excel._load_mappings()
    assert result == {}


def test_load_mappings_reads_toml(tmp_path, monkeypatch):
    toml_content = '[mappings]\n"TestCo" = "Groceries"\n'
    f = tmp_path / 'mappings.toml'
    f.write_text(toml_content, encoding='utf-8')
    monkeypatch.setattr(fill_excel, '_MAPPINGS_FILE', f)
    result = fill_excel._load_mappings()
    assert result['TestCo'] == 'Groceries'


def test_load_mappings_handles_empty_mappings_section(tmp_path, monkeypatch):
    toml_content = '[mappings]\n'
    f = tmp_path / 'mappings.toml'
    f.write_text(toml_content, encoding='utf-8')
    monkeypatch.setattr(fill_excel, '_MAPPINGS_FILE', f)
    assert fill_excel._load_mappings() == {}
